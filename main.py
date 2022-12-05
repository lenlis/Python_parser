import csv
import sys
import openpyxl as pxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

all_salery_by_year = {}

num_by_year = {}

profession_salery_by_year = {}

profession_count_by_year = {}

salery_by_city = {}

share_by_city = {}

sorted_salery_by_city = {}

sorted_share_by_city = {}

years =[]

years_exel_rows = []

first_city_exel_rows = []

second_city_exel_rows = []

thin = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

#vacancies_by_year.csv
class Report:
    @staticmethod
    def generate_excel():
        prof_sal_lable = "Средняя зарплата - " + start.prof_name
        prof_count_lable = "Количество вакансий - " + start.prof_name
        year_data = [[years, 'A', years, "Год"], [all_salery_by_year.values(), 'B', years, "Средняя зарплата"], [num_by_year.values(), 'D', years, "Количество вакансий"], [profession_salery_by_year.values(), 'C', years, prof_sal_lable], [profession_count_by_year.values(), 'E', years, prof_count_lable]]
        city_data = [[sorted_salery_by_city.keys(), 'A', sorted_salery_by_city.keys(), "Город"], [sorted_salery_by_city.values(), 'B', sorted_salery_by_city.keys(), "Уровень зарплат"], [sorted_share_by_city.keys(), 'D', sorted_share_by_city.keys(), "Город"], [sorted_share_by_city.values(), 'E', sorted_share_by_city.keys(), "Доля вакансий"]]
        wb = pxl.Workbook()

        wb.active.title = "Статистика по годам"

        wb.create_sheet("Статистика по городам")

        ws1 = wb["Статистика по годам"]
        ws2 = wb["Статистика по городам"]

        Report.name_and_fill_cols(ws1, year_data)
        Report.name_and_fill_cols(ws2, city_data)
        Report.fill_column(year_data, ws1)
        Report.fill_column(city_data, ws2)

        Report.generate_rows_for_pdf(ws1, len(years)+1, ['A', 'B', 'C', 'D', 'E'], years_exel_rows)
        Report.generate_rows_for_pdf(ws2, 12, ['A', 'B'], first_city_exel_rows)
        Report.generate_rows_for_pdf(ws2, 12, ['D', 'E'], second_city_exel_rows)

        wb.save("report.xlsx")

    @staticmethod
    def count_col_width(col):
        column_width = 0
        data = []
        for v in col[0]:
            data.append(v)
        data.append(col[3])
        for item in data:
            if len(str(item)) > column_width:
                column_width = len(str(item))
        return column_width

    @staticmethod
    def fill_column(data, list):
        for values in data:
            row = 2
            for value in values[0]:
                col = values[1]
                if values[3] == "Год":
                    list[col + str(row)] = int(value)
                elif values[3] == "Доля вакансий":
                    list[col + str(row)] = str(float(value)*100)[:4] + '%'
                else:
                    list[col + str(row)] = value
                list[col + str(row)].border = thin
                row += 1

    @staticmethod
    def name_and_fill_cols(ws, cols):
        for col in cols:
            ws[col[1] + '1'] = col[3]
            ws[col[1] + '1'].font = Font(bold=True)
            ws[col[1] + '1'].border = thin
            ws.column_dimensions[col[1]].width = Report.count_col_width(col)

    @staticmethod
    def generate_image():
        matplotlib.rcParams.update({'font.size': 8})
        fig, axs = plt.subplots(2, 2)
        Report.create_bar_diag(axs[0, 0], all_salery_by_year, profession_salery_by_year, years, 'Уровень зарплат по годам', 'средняя з/п', f'з/п {start.prof_name.lower()}')
        Report.create_bar_diag(axs[0, 1], num_by_year, profession_count_by_year, years, 'Количество вакансий по годам', 'Количество вакансий', f'Количество вакансий \n{start.prof_name.lower()}')
        Report.create_hor_bar_diag(axs[1, 0], sorted_salery_by_city, 'Уровень зарплат по городам')
        Report.create_pie_diag(axs[1, 1], sorted_share_by_city, share_by_city, 'Доля вакансий по городам')
        fig.tight_layout()
        plt.savefig("D:\\PythonProject\\report\\graph.png")

    @staticmethod
    def create_bar_diag(gr, first, sec, hor, tit, l1, l2):
        first_res = []
        sec_res = []
        X_axis = np.arange(len(hor))
        for v in first.values():
            first_res.append(v)
        for v in sec.values():
            sec_res.append(v)
        gr.bar(X_axis - 0.2, first_res, 0.4, label=l1)
        gr.bar(X_axis + 0.2, sec_res, 0.4, label=l2)
        gr.set_xticks(X_axis, hor, rotation=90)
        gr.set_title(tit)
        gr.legend(loc='upper left')

    @staticmethod
    def create_hor_bar_diag(gr, data, tit):
        first_res = []
        vert = []
        Y_axis = np.arange(len(data))
        for v in data.values():
            first_res.append(v)
        first_res.reverse()
        for v in data.keys():
            vert.append(v)
        vert.reverse()
        gr.barh(Y_axis, first_res, 0.4)
        gr.set_yticks(Y_axis, vert, size=6)
        gr.set_title(tit)

    @staticmethod
    def create_pie_diag(gr, data, not_sorted_data, tit):
        values = []
        lables = []
        others = 0
        lables.append('Другие')
        for key in not_sorted_data.keys():
            if key not in data.keys():
                others += not_sorted_data[key]
        values.append(others)
        for v in data.values():
            values.append(v)
        for v in data.keys():
            lables.append(v)
        gr.pie(values, labels=lables, textprops={'fontsize': 6})
        gr.set_title(tit)

    @staticmethod
    def generate_rows_for_pdf(ws, border, cols, rows_list):
        for i in range(2,border):
            row = []
            for col in cols:
                row.append(ws[col + str(i)].value)
            rows_list.append(row)

    @staticmethod
    def generate_pdf():
        path = Environment(loader=FileSystemLoader("."))
        template = path.get_template("pattern.html")
        pdf_template = template.render(prof = start.prof_name, year_rows = years_exel_rows, first_city_rows = first_city_exel_rows, second_city_rows = second_city_exel_rows)
        config = pdfkit.configuration(wkhtmltopdf="D:\\wkhtmltopdf\\bin\\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

class DataSet:
    vacancies_objects = []
    @staticmethod
    def csv_reader(file_name):
        with open(file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            vacancies_list = []
            file_len = len(file.readlines())
            if file_len == 0:
                print("Пустой файл")
                sys.exit()
            elif file_len == 1:
                print("Нет данных")
                sys.exit()
            file.seek(0)
            csv_headers = reader.__next__()
            csv_rows = [x for x in reader]
            for vacancy in csv_rows:
                dict = {}
                if len(csv_headers) == len(vacancy) and vacancy.count('') == 0:
                    for i in range(len(vacancy)):
                        dict[csv_headers[i]] = vacancy[i]
                    vacancies_list.append(dict)
            return vacancies_list

    @staticmethod
    def csv_filer(vacancies_list):
        countVacancies = 0
        data_vacancies = DataSet()
        data_vacancies.vacancies_objects = []
        for vac in vacancies_list:
            data_vacancies, countVacancies = DataSet.formatter(vac, data_vacancies, countVacancies)

        for i in all_salery_by_year:
            all_salery_by_year[i] = int(
                all_salery_by_year[i][0] / all_salery_by_year[i][1])

        for i in profession_salery_by_year:
            profession_salery_by_year[i] = int(
                profession_salery_by_year[i][0] /
                profession_salery_by_year[i][1])

        count = 0
        for i in share_by_city:
            share_by_city[i] = round(int(share_by_city[i]) / countVacancies, 4)
        sortedVacanciesKeys = sorted(share_by_city, key=share_by_city.get, reverse=True)
        for w in sortedVacanciesKeys:
            if (count >= 10 or share_by_city[w] < 0.01):
                break
            count += 1
            sorted_share_by_city[w] = share_by_city[w]

        count = 0
        for i in salery_by_city:
            salery_by_city[i] = int(salery_by_city[i][0] / salery_by_city[i][1])
        sortedSalaryKeys = sorted(salery_by_city, key=salery_by_city.get, reverse=True)

        for w in sortedSalaryKeys:
            if share_by_city[w] >= 0.01:
                if (count >= 10):
                    break
                count += 1
                sorted_salery_by_city[w] = salery_by_city[w]

        DataSet.year_check(years)
        start.print_vacancies()

    @staticmethod
    def formatter(vac, data_set, count):
        DataSet.add_year(vac['published_at'][:4])
        DataSet.complet_dictionary(all_salery_by_year, vac, vac['published_at'][:4])
        DataSet.complet_count_dictionary(num_by_year, vac['published_at'][:4])
        DataSet.complet_dictionary(salery_by_city, vac, vac['area_name'])
        DataSet.complet_count_dictionary(share_by_city, vac['area_name'])

        if start.prof_name in vac['name']:
            DataSet.complet_dictionary(profession_salery_by_year, vac, vac['published_at'][:4])
            DataSet.complet_count_dictionary(profession_count_by_year, vac['published_at'][:4])
        count += 1
        return data_set, count

    @staticmethod
    def year_check(years):
        for year in years:
            if year not in profession_salery_by_year.keys():
                profession_salery_by_year[year] = 0

            if year not in profession_count_by_year.keys():
                profession_count_by_year[year] = 0

    @staticmethod
    def complet_dictionary(dictCompleting,dict,key):
        if key not in dictCompleting.keys():
            dictCompleting[key] = [
                int((float(dict['salary_from']) + float(dict['salary_to'])) / 2) * valuta_description[dict['salary_currency']][1], 1]
        else:
            dictCompleting[key] = [
                int(dictCompleting[key][0]) + int(
                    (float(dict['salary_from']) + float(dict['salary_to'])) * valuta_description[dict['salary_currency']][1] / 2),
                int(dictCompleting[key][1]) + 1]

    @staticmethod
    def complet_count_dictionary(dictCompleting,key):
        if key not in dictCompleting.keys():
            dictCompleting[key] = 1
        else:
            dictCompleting[key] = int(dictCompleting[key]) + 1

    @staticmethod
    def add_year(year):
        if year not in years:
            years.append(year)


valuta_description = {
    'AZN': ['Манаты', 35.68],
    'BYR': ['Белорусские рубли', 23.91],
    'EUR': ['Евро', 59.90],
    'GEL': ['Грузинский лари', 21.74],
    'KGS': ['Киргизский сом', 0.76],
    'KZT': ['Тенге', 0.13],
    'RUR': ['Рубли', 1],
    'UAH': ['Гривны', 1.64],
    'USD': ['Доллары', 60.66],
    'UZS': ['Узбекский сум', 0.0055]
}


class InputConect:
    name = ''
    prof_name = ''
    command = ''

    def input(self):
        self.name = input("Введите название файла: ")
        self.prof_name = input("Введите название профессии: ")
        self.command = input("Введите команду вывода(Вакансии или Статистика): ")
        if self.command not in ['Вакансии', 'Статистика']:
            print('invalid command')
            sys.exit()

    def print_vacancies(self):
        print("Динамика уровня зарплат по годам: {0}".format(all_salery_by_year).replace("'", ''))
        print(
            "Динамика количества вакансий по годам: {0}".format(num_by_year).replace("'", ''))
        print("Динамика уровня зарплат по годам для выбранной профессии: {0}".format(
            profession_salery_by_year).replace("'", ''))
        print("Динамика количества вакансий по годам для выбранной профессии: {0}".format(
            profession_count_by_year).replace("'", ''))
        print("Уровень зарплат по городам (в порядке убывания): {0}".format(sorted_salery_by_city))
        print("Доля вакансий по городам (в порядке убывания): {0}".format(sorted_share_by_city))
        if self.command == 'Вакансии':
            Report.generate_excel()
        else:
            Report.generate_image()
        # Report.generate_pdf()


start = InputConect()
start.input()
vacancies_list = DataSet.csv_reader(start.name)
DataSet.csv_filer(vacancies_list)

